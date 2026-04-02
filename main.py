import sys
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font as XlFont, Alignment, PatternFill, Border, Side
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTabWidget, QTableWidget, QTableWidgetItem, QPushButton, QLabel,
    QLineEdit, QSpinBox, QDateEdit, QDialog, QFormLayout, QMessageBox,
    QComboBox, QDoubleSpinBox, QGroupBox, QSplitter, QTextEdit,
    QHeaderView, QAbstractItemView, QGridLayout, QScrollArea,
    QDialogButtonBox, QSizePolicy, QFileDialog, QFrame
)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QFont, QColor

from database import Database


def fmt_money(amount):
    """Định dạng tiền tệ VND"""
    try:
        return f"{int(amount):,}đ".replace(",", ".")
    except (ValueError, TypeError):
        return "0đ"


def fmt_month_display(month_str):
    """Chuyển YYYY-MM sang M/YYYY"""
    try:
        y, m = month_str.split("-")
        return f"{int(m)}/{y}"
    except (ValueError, AttributeError):
        return month_str


class RoomDialog(QDialog):
    """Dialog thêm/sửa phòng"""
    def __init__(self, parent=None, room_data=None):
        super().__init__(parent)
        self.room_data = room_data
        self.setWindowTitle("Sửa Phòng" if room_data else "Thêm Phòng")
        self.setMinimumWidth(380)
        self._build_ui()

    def _build_ui(self):
        layout = QFormLayout(self)
        layout.setSpacing(10)

        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("VD: P101")

        self.price_input = QDoubleSpinBox()
        self.price_input.setMaximum(99_000_000)
        self.price_input.setSingleStep(100_000)
        self.price_input.setDecimals(0)
        self.price_input.setSuffix(" đ")

        self.start_date_input = QDateEdit()
        self.start_date_input.setCalendarPopup(True)
        self.start_date_input.setDate(QDate.currentDate())
        self.start_date_input.setDisplayFormat("dd/MM/yyyy")

        self.deposit_input = QDoubleSpinBox()
        self.deposit_input.setMaximum(99_000_000)
        self.deposit_input.setSingleStep(100_000)
        self.deposit_input.setDecimals(0)
        self.deposit_input.setSuffix(" đ")

        self.billing_day_input = QSpinBox()
        self.billing_day_input.setRange(1, 28)
        self.billing_day_input.setValue(1)

        self.notes_input = QLineEdit()
        self.notes_input.setPlaceholderText("Ghi chú (tuỳ chọn)")

        if self.room_data:
            self.name_input.setText(self.room_data["name"])
            self.price_input.setValue(self.room_data["price"])
            if self.room_data["start_date"]:
                try:
                    d = QDate.fromString(self.room_data["start_date"], "yyyy-MM-dd")
                    self.start_date_input.setDate(d)
                except Exception:
                    pass
            self.deposit_input.setValue(self.room_data["deposit"] or 0)
            self.billing_day_input.setValue(self.room_data["billing_day"] or 1)
            self.notes_input.setText(self.room_data["notes"] or "")

        # Auto-fill billing day from start date
        self.start_date_input.dateChanged.connect(self._on_start_date_changed)

        layout.addRow("Tên Phòng *:", self.name_input)
        layout.addRow("Giá Phòng (đ/tháng) *:", self.price_input)
        layout.addRow("Ngày Bắt Đầu Ở:", self.start_date_input)
        layout.addRow("Đặt Cọc:", self.deposit_input)
        layout.addRow("Ngày Tính Tiền (ngày):", self.billing_day_input)
        layout.addRow("Ghi Chú:", self.notes_input)

        btn_box = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        btn_box.accepted.connect(self._validate_and_accept)
        btn_box.rejected.connect(self.reject)
        layout.addRow(btn_box)

    def _on_start_date_changed(self, qdate):
        self.billing_day_input.setValue(qdate.day())

    def _validate_and_accept(self):
        if not self.name_input.text().strip():
            QMessageBox.warning(self, "Lỗi", "Vui lòng nhập tên phòng!")
            return
        if self.price_input.value() <= 0:
            QMessageBox.warning(self, "Lỗi", "Vui lòng nhập giá phòng!")
            return
        self.accept()

    def get_data(self):
        return {
            "name": self.name_input.text().strip(),
            "price": self.price_input.value(),
            "start_date": self.start_date_input.date().toString("yyyy-MM-dd"),
            "deposit": self.deposit_input.value(),
            "billing_day": self.billing_day_input.value(),
            "notes": self.notes_input.text().strip(),
        }


class TenantDialog(QDialog):
    """Dialog thêm/sửa khách trọ"""
    def __init__(self, parent=None, tenant_data=None):
        super().__init__(parent)
        self.tenant_data = tenant_data
        self.setWindowTitle("Sửa Khách Trọ" if tenant_data else "Thêm Khách Trọ")
        self.setMinimumWidth(380)
        self._build_ui()

    def _build_ui(self):
        layout = QFormLayout(self)
        layout.setSpacing(10)

        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("Họ và tên *")

        self.cccd_input = QLineEdit()
        self.cccd_input.setPlaceholderText("Số CCCD (tuỳ chọn)")

        self.dob_input = QLineEdit()
        self.dob_input.setPlaceholderText("VD: 15/05/1990 (tuỳ chọn)")

        self.hometown_input = QLineEdit()
        self.hometown_input.setPlaceholderText("Quê quán (tuỳ chọn)")

        self.phone_input = QLineEdit()
        self.phone_input.setPlaceholderText("Số điện thoại (tuỳ chọn)")

        if self.tenant_data:
            self.name_input.setText(self.tenant_data["name"] or "")
            self.cccd_input.setText(self.tenant_data["cccd"] or "")
            self.dob_input.setText(self.tenant_data["dob"] or "")
            self.hometown_input.setText(self.tenant_data["hometown"] or "")
            self.phone_input.setText(self.tenant_data["phone"] or "")

        layout.addRow("Tên Khách Trọ *:", self.name_input)
        layout.addRow("CCCD:", self.cccd_input)
        layout.addRow("Ngày Tháng Năm Sinh:", self.dob_input)
        layout.addRow("Quê Quán:", self.hometown_input)
        layout.addRow("Số Điện Thoại:", self.phone_input)

        btn_box = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        btn_box.accepted.connect(self._validate_and_accept)
        btn_box.rejected.connect(self.reject)
        layout.addRow(btn_box)

    def _validate_and_accept(self):
        if not self.name_input.text().strip():
            QMessageBox.warning(self, "Lỗi", "Vui lòng nhập tên khách trọ!")
            return
        self.accept()

    def get_data(self):
        return {
            "name": self.name_input.text().strip(),
            "cccd": self.cccd_input.text().strip(),
            "dob": self.dob_input.text().strip(),
            "hometown": self.hometown_input.text().strip(),
            "phone": self.phone_input.text().strip(),
        }


class RoomManagementApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.db = Database()
        self.selected_room_id = None
        self.init_ui()
        self._check_billing_notifications()

    def init_ui(self):
        self.setWindowTitle("🏠 Quản Lý Phòng Trọ")
        self.setGeometry(100, 100, 1100, 700)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(8, 8, 8, 8)
        main_layout.setSpacing(6)

        # Tiêu đề
        title = QLabel("🏠 HỆ THỐNG QUẢN LÝ PHÒNG TRỌ")
        title_font = QFont()
        title_font.setPointSize(15)
        title_font.setBold(True)
        title.setFont(title_font)
        title.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title)

        # Dashboard
        dash_frame = QFrame()
        dash_frame.setFrameShape(QFrame.StyledPanel)
        dash_layout = QHBoxLayout(dash_frame)
        self.lbl_total_rooms = QLabel("Tổng phòng: 0")
        self.lbl_empty_rooms = QLabel("Phòng trống: 0")
        self.lbl_total_tenants = QLabel("Tổng khách trọ: 0")
        for lbl in [self.lbl_total_rooms, self.lbl_empty_rooms, self.lbl_total_tenants]:
            lbl.setAlignment(Qt.AlignCenter)
            f = QFont()
            f.setPointSize(10)
            f.setBold(True)
            lbl.setFont(f)
            dash_layout.addWidget(lbl)
        main_layout.addWidget(dash_frame)

        # Tabs
        tabs = QTabWidget()
        tabs.addTab(self._create_rooms_tab(), "🏠 Phòng")
        tabs.addTab(self._create_utility_tab(), "⚡ Điện Nước")
        tabs.addTab(self._create_billing_tab(), "📋 Tính Tiền")
        tabs.addTab(self._create_export_tab(), "📊 Xuất Excel")
        tabs.addTab(self._create_settings_tab(), "⚙️ Cài Đặt")
        main_layout.addWidget(tabs)

        self._refresh_dashboard()

    # ==================== TAB PHÒNG ====================
    def _create_rooms_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # --- Phần quản lý phòng ---
        room_group = QGroupBox("📋 Danh Sách Phòng")
        room_v = QVBoxLayout(room_group)

        btn_row = QHBoxLayout()
        btn_add_room = QPushButton("➕ Thêm Phòng")
        btn_add_room.clicked.connect(self._add_room)
        btn_edit_room = QPushButton("✏️ Sửa Phòng")
        btn_edit_room.clicked.connect(self._edit_room)
        btn_del_room = QPushButton("🗑️ Xóa Phòng")
        btn_del_room.clicked.connect(self._delete_room)
        btn_row.addWidget(btn_add_room)
        btn_row.addWidget(btn_edit_room)
        btn_row.addWidget(btn_del_room)
        btn_row.addStretch()
        room_v.addLayout(btn_row)

        self.rooms_table = QTableWidget()
        self.rooms_table.setColumnCount(7)
        self.rooms_table.setHorizontalHeaderLabels([
            "Tên Phòng", "Giá/Tháng", "Ngày BĐ Ở", "Đặt Cọc",
            "Ngày TT", "Trạng Thái", "Ghi Chú"
        ])
        self.rooms_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.rooms_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.rooms_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.rooms_table.setAlternatingRowColors(True)
        self.rooms_table.itemSelectionChanged.connect(self._on_room_selected)
        room_v.addWidget(self.rooms_table)
        layout.addWidget(room_group, 3)

        # --- Phần quản lý khách trọ ---
        tenant_group = QGroupBox("👥 Khách Trọ Trong Phòng")
        tenant_v = QVBoxLayout(tenant_group)

        self.lbl_selected_room = QLabel("← Chọn phòng ở trên để xem khách trọ")
        self.lbl_selected_room.setStyleSheet("color: gray; font-style: italic;")
        tenant_v.addWidget(self.lbl_selected_room)

        btn_row2 = QHBoxLayout()
        btn_add_tenant = QPushButton("➕ Thêm Khách Trọ")
        btn_add_tenant.clicked.connect(self._add_tenant)
        btn_edit_tenant = QPushButton("✏️ Sửa Khách Trọ")
        btn_edit_tenant.clicked.connect(self._edit_tenant)
        btn_del_tenant = QPushButton("🗑️ Xóa Khách Trọ")
        btn_del_tenant.clicked.connect(self._delete_tenant)
        btn_row2.addWidget(btn_add_tenant)
        btn_row2.addWidget(btn_edit_tenant)
        btn_row2.addWidget(btn_del_tenant)
        btn_row2.addStretch()
        tenant_v.addLayout(btn_row2)

        self.tenants_table = QTableWidget()
        self.tenants_table.setColumnCount(5)
        self.tenants_table.setHorizontalHeaderLabels([
            "Tên Khách Trọ", "CCCD", "Ngày Sinh", "Quê Quán", "Số Điện Thoại"
        ])
        self.tenants_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tenants_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tenants_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tenants_table.setAlternatingRowColors(True)
        tenant_v.addWidget(self.tenants_table)
        layout.addWidget(tenant_group, 2)

        self._refresh_rooms_table()
        return widget

    def _refresh_rooms_table(self):
        rooms = self.db.get_all_rooms()
        self.rooms_table.setRowCount(len(rooms))
        for row, r in enumerate(rooms):
            status_text = "Có người ✅" if r["status"] == "occupied" else "Trống 🔵"
            start_date_display = ""
            if r["start_date"]:
                try:
                    d = datetime.strptime(r["start_date"], "%Y-%m-%d")
                    start_date_display = d.strftime("%d/%m/%Y")
                except ValueError:
                    start_date_display = r["start_date"]
            items = [
                r["name"],
                fmt_money(r["price"]),
                start_date_display,
                fmt_money(r["deposit"] or 0),
                f"Ngày {r['billing_day']}",
                status_text,
                r["notes"] or "",
            ]
            for col, val in enumerate(items):
                item = QTableWidgetItem(str(val))
                item.setData(Qt.UserRole, r["id"])
                if r["status"] == "occupied":
                    item.setBackground(QColor("#e8f5e9"))
                self.rooms_table.setItem(row, col, item)
        self._refresh_dashboard()

    def _on_room_selected(self):
        rows = self.rooms_table.selectedItems()
        if not rows:
            self.selected_room_id = None
            self.lbl_selected_room.setText("← Chọn phòng ở trên để xem khách trọ")
            self.tenants_table.setRowCount(0)
            return
        room_id = self.rooms_table.item(self.rooms_table.currentRow(), 0).data(Qt.UserRole)
        self.selected_room_id = room_id
        room = self.db.get_room(room_id)
        self.lbl_selected_room.setText(f"Khách trọ trong phòng: {room['name']}")
        self._refresh_tenants_table(room_id)

    def _refresh_tenants_table(self, room_id):
        tenants = self.db.get_tenants_by_room(room_id)
        self.tenants_table.setRowCount(len(tenants))
        for row, t in enumerate(tenants):
            items = [
                t["name"],
                t["cccd"] or "",
                t["dob"] or "",
                t["hometown"] or "",
                t["phone"] or "",
            ]
            for col, val in enumerate(items):
                item = QTableWidgetItem(val)
                item.setData(Qt.UserRole, t["id"])
                self.tenants_table.setItem(row, col, item)

    def _add_room(self):
        dlg = RoomDialog(self)
        if dlg.exec_() == QDialog.Accepted:
            data = dlg.get_data()
            try:
                self.db.add_room(
                    data["name"], data["price"], data["start_date"],
                    data["deposit"], data["billing_day"], data["notes"]
                )
                self._refresh_rooms_table()
                QMessageBox.information(self, "✅ Thành công", f"Đã thêm phòng {data['name']}!")
            except Exception as e:
                QMessageBox.critical(self, "❌ Lỗi", f"Không thể thêm phòng:\n{e}")

    def _edit_room(self):
        if self.selected_room_id is None:
            QMessageBox.warning(self, "Thông báo", "Vui lòng chọn phòng cần sửa!")
            return
        room = self.db.get_room(self.selected_room_id)
        dlg = RoomDialog(self, dict(room))
        if dlg.exec_() == QDialog.Accepted:
            data = dlg.get_data()
            try:
                self.db.update_room(
                    self.selected_room_id, data["name"], data["price"],
                    data["start_date"], data["deposit"], data["billing_day"], data["notes"]
                )
                self._refresh_rooms_table()
                QMessageBox.information(self, "✅ Thành công", "Đã cập nhật thông tin phòng!")
            except Exception as e:
                QMessageBox.critical(self, "❌ Lỗi", f"Không thể cập nhật phòng:\n{e}")

    def _delete_room(self):
        if self.selected_room_id is None:
            QMessageBox.warning(self, "Thông báo", "Vui lòng chọn phòng cần xóa!")
            return
        room = self.db.get_room(self.selected_room_id)
        reply = QMessageBox.question(
            self, "Xác nhận xóa",
            f"Bạn có chắc muốn xóa phòng {room['name']}?\n(Tất cả khách trọ trong phòng cũng sẽ bị xóa)",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            self.db.delete_room(self.selected_room_id)
            self.selected_room_id = None
            self.tenants_table.setRowCount(0)
            self.lbl_selected_room.setText("← Chọn phòng ở trên để xem khách trọ")
            self._refresh_rooms_table()

    def _add_tenant(self):
        if self.selected_room_id is None:
            QMessageBox.warning(self, "Thông báo", "Vui lòng chọn phòng trước khi thêm khách trọ!")
            return
        dlg = TenantDialog(self)
        if dlg.exec_() == QDialog.Accepted:
            data = dlg.get_data()
            try:
                self.db.add_tenant(
                    self.selected_room_id, data["name"], data["cccd"],
                    data["dob"], data["hometown"], data["phone"]
                )
                self._refresh_tenants_table(self.selected_room_id)
                self._refresh_rooms_table()
                # Re-select the same room
                self._reselect_room(self.selected_room_id)
                QMessageBox.information(self, "✅ Thành công", f"Đã thêm khách trọ {data['name']}!")
            except Exception as e:
                QMessageBox.critical(self, "❌ Lỗi", f"Không thể thêm khách trọ:\n{e}")

    def _edit_tenant(self):
        row = self.tenants_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Thông báo", "Vui lòng chọn khách trọ cần sửa!")
            return
        tenant_id = self.tenants_table.item(row, 0).data(Qt.UserRole)
        tenant = self.db.get_tenant(tenant_id)
        dlg = TenantDialog(self, dict(tenant))
        if dlg.exec_() == QDialog.Accepted:
            data = dlg.get_data()
            try:
                self.db.update_tenant(
                    tenant_id, data["name"], data["cccd"],
                    data["dob"], data["hometown"], data["phone"]
                )
                self._refresh_tenants_table(self.selected_room_id)
                QMessageBox.information(self, "✅ Thành công", "Đã cập nhật thông tin khách trọ!")
            except Exception as e:
                QMessageBox.critical(self, "❌ Lỗi", f"Không thể cập nhật:\n{e}")

    def _delete_tenant(self):
        row = self.tenants_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Thông báo", "Vui lòng chọn khách trọ cần xóa!")
            return
        tenant_id = self.tenants_table.item(row, 0).data(Qt.UserRole)
        name = self.tenants_table.item(row, 0).text()
        reply = QMessageBox.question(
            self, "Xác nhận xóa",
            f"Bạn có chắc muốn xóa khách trọ '{name}'?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            self.db.delete_tenant(tenant_id)
            self._refresh_tenants_table(self.selected_room_id)
            self._refresh_rooms_table()
            self._reselect_room(self.selected_room_id)

    def _reselect_room(self, room_id):
        """Chọn lại phòng trong bảng sau khi refresh"""
        for row in range(self.rooms_table.rowCount()):
            item = self.rooms_table.item(row, 0)
            if item and item.data(Qt.UserRole) == room_id:
                self.rooms_table.selectRow(row)
                break

    # ==================== TAB ĐIỆN NƯỚC ====================
    def _create_utility_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # Row: Room selector + Month selector
        selector_row = QHBoxLayout()
        selector_row.addWidget(QLabel("Phòng:"))
        self.util_room_combo = QComboBox()
        self.util_room_combo.setMinimumWidth(120)
        selector_row.addWidget(self.util_room_combo)
        selector_row.addSpacing(20)
        selector_row.addWidget(QLabel("Tháng:"))
        self.util_month_combo = QComboBox()
        self._populate_month_combo(self.util_month_combo)
        selector_row.addWidget(self.util_month_combo)
        btn_load_util = QPushButton("🔍 Tải Dữ Liệu")
        btn_load_util.clicked.connect(self._load_utility_data)
        selector_row.addWidget(btn_load_util)
        selector_row.addStretch()
        layout.addLayout(selector_row)

        # Grid for electricity + water forms
        forms_layout = QHBoxLayout()

        # Electricity group
        elec_group = QGroupBox("⚡ Điện")
        elec_form = QFormLayout(elec_group)
        self.elec_old = QDoubleSpinBox()
        self.elec_old.setMaximum(9_999_999)
        self.elec_old.setDecimals(0)
        self.elec_new = QDoubleSpinBox()
        self.elec_new.setMaximum(9_999_999)
        self.elec_new.setDecimals(0)
        self.elec_usage_lbl = QLabel("Xài: 0 kWh")
        self.elec_old.valueChanged.connect(self._calc_elec_usage)
        self.elec_new.valueChanged.connect(self._calc_elec_usage)
        elec_form.addRow("Số Cũ (kWh):", self.elec_old)
        elec_form.addRow("Số Mới (kWh):", self.elec_new)
        elec_form.addRow("", self.elec_usage_lbl)
        btn_save_elec = QPushButton("💾 Lưu Điện")
        btn_save_elec.clicked.connect(self._save_electricity)
        elec_form.addRow(btn_save_elec)
        forms_layout.addWidget(elec_group)

        # Water group
        water_group = QGroupBox("💧 Nước")
        water_form = QFormLayout(water_group)
        self.water_old = QDoubleSpinBox()
        self.water_old.setMaximum(9_999_999)
        self.water_old.setDecimals(0)
        self.water_new = QDoubleSpinBox()
        self.water_new.setMaximum(9_999_999)
        self.water_new.setDecimals(0)
        self.water_usage_lbl = QLabel("Xài: 0 khối")
        self.water_old.valueChanged.connect(self._calc_water_usage)
        self.water_new.valueChanged.connect(self._calc_water_usage)
        water_form.addRow("Số Cũ (khối):", self.water_old)
        water_form.addRow("Số Mới (khối):", self.water_new)
        water_form.addRow("", self.water_usage_lbl)
        btn_save_water = QPushButton("💾 Lưu Nước")
        btn_save_water.clicked.connect(self._save_water)
        water_form.addRow(btn_save_water)
        forms_layout.addWidget(water_group)

        layout.addLayout(forms_layout)

        # History table
        hist_group = QGroupBox("📜 Lịch Sử Số Điện")
        hist_v = QVBoxLayout(hist_group)
        self.elec_history_table = QTableWidget()
        self.elec_history_table.setColumnCount(5)
        self.elec_history_table.setHorizontalHeaderLabels(["Tháng", "Số Cũ", "Số Mới", "Xài (kWh)", "Thành Tiền"])
        self.elec_history_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.elec_history_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.elec_history_table.setAlternatingRowColors(True)
        hist_v.addWidget(self.elec_history_table)
        layout.addWidget(hist_group)

        self._refresh_util_room_combo()
        return widget

    def _populate_month_combo(self, combo):
        combo.clear()
        now = datetime.now()
        for i in range(12):
            m = now.month - i
            y = now.year
            while m <= 0:
                m += 12
                y -= 1
            month_str = f"{y}-{m:02d}"
            combo.addItem(fmt_month_display(month_str), month_str)

    def _refresh_util_room_combo(self):
        self.util_room_combo.clear()
        rooms = self.db.get_all_rooms()
        for r in rooms:
            self.util_room_combo.addItem(r["name"], r["id"])

    def _get_selected_util_room_id(self):
        return self.util_room_combo.currentData()

    def _get_selected_util_month(self):
        return self.util_month_combo.currentData()

    def _calc_elec_usage(self):
        usage = max(0, self.elec_new.value() - self.elec_old.value())
        price = float(self.db.get_setting("electricity_price", 3500))
        fee = usage * price
        self.elec_usage_lbl.setText(f"Xài: {usage:.0f} kWh  →  {fmt_money(fee)}")

    def _calc_water_usage(self):
        usage = max(0, self.water_new.value() - self.water_old.value())
        price = float(self.db.get_setting("water_price", 15000))
        fee = usage * price
        self.water_usage_lbl.setText(f"Xài: {usage:.0f} khối  →  {fmt_money(fee)}")

    def _load_utility_data(self):
        room_id = self._get_selected_util_room_id()
        month = self._get_selected_util_month()
        if not room_id:
            return

        elec = self.db.get_electricity(room_id, month)
        if elec:
            self.elec_old.setValue(elec["old_reading"])
            self.elec_new.setValue(elec["new_reading"])
        else:
            self.elec_old.setValue(0)
            self.elec_new.setValue(0)

        water = self.db.get_water(room_id, month)
        if water:
            self.water_old.setValue(water["old_reading"])
            self.water_new.setValue(water["new_reading"])
        else:
            self.water_old.setValue(0)
            self.water_new.setValue(0)

        # Load electricity history
        history = self.db.get_electricity_history(room_id)
        price = float(self.db.get_setting("electricity_price", 3500))
        self.elec_history_table.setRowCount(len(history))
        for row, rec in enumerate(history):
            usage = max(0, rec["new_reading"] - rec["old_reading"])
            fee = usage * price
            vals = [
                fmt_month_display(rec["month"]),
                f"{rec['old_reading']:.0f}",
                f"{rec['new_reading']:.0f}",
                f"{usage:.0f}",
                fmt_money(fee),
            ]
            for col, v in enumerate(vals):
                self.elec_history_table.setItem(row, col, QTableWidgetItem(v))

    def _save_electricity(self):
        room_id = self._get_selected_util_room_id()
        month = self._get_selected_util_month()
        if not room_id:
            QMessageBox.warning(self, "Thông báo", "Vui lòng chọn phòng!")
            return
        old_v = self.elec_old.value()
        new_v = self.elec_new.value()
        if new_v < old_v:
            QMessageBox.warning(self, "Lỗi", "Số mới không thể nhỏ hơn số cũ!")
            return
        self.db.save_electricity(room_id, month, old_v, new_v)
        QMessageBox.information(self, "✅ Thành công", f"Đã lưu số điện tháng {fmt_month_display(month)}!")
        self._load_utility_data()

    def _save_water(self):
        room_id = self._get_selected_util_room_id()
        month = self._get_selected_util_month()
        if not room_id:
            QMessageBox.warning(self, "Thông báo", "Vui lòng chọn phòng!")
            return
        old_v = self.water_old.value()
        new_v = self.water_new.value()
        if new_v < old_v:
            QMessageBox.warning(self, "Lỗi", "Số mới không thể nhỏ hơn số cũ!")
            return
        self.db.save_water(room_id, month, old_v, new_v)
        QMessageBox.information(self, "✅ Thành công", f"Đã lưu số nước tháng {fmt_month_display(month)}!")

    # ==================== TAB TÍNH TIỀN ====================
    def _create_billing_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # Selector row
        sel_row = QHBoxLayout()
        sel_row.addWidget(QLabel("Phòng:"))
        self.bill_room_combo = QComboBox()
        self.bill_room_combo.setMinimumWidth(120)
        sel_row.addWidget(self.bill_room_combo)
        sel_row.addSpacing(20)
        sel_row.addWidget(QLabel("Tháng:"))
        self.bill_month_combo = QComboBox()
        self._populate_month_combo(self.bill_month_combo)
        sel_row.addWidget(self.bill_month_combo)
        btn_calc = QPushButton("🧮 Tính Hóa Đơn")
        btn_calc.clicked.connect(self._calculate_bill)
        sel_row.addWidget(btn_calc)
        sel_row.addStretch()
        layout.addLayout(sel_row)

        # Main area: left = inputs, right = bill preview
        content = QHBoxLayout()

        # Left: editable fields
        inputs_group = QGroupBox("✏️ Nhập Chi Phí")
        inputs_form = QFormLayout(inputs_group)

        self.bill_room_fee_lbl = QLabel("0đ")
        self.bill_elec_lbl = QLabel("0đ  (chưa nhập số điện)")

        # Water
        self.bill_water_lbl = QLabel("0đ  (chưa nhập số nước)")

        # Laundry
        self.bill_laundry_times = QSpinBox()
        self.bill_laundry_times.setRange(0, 99)
        self.bill_laundry_times.setPrefix("  ")
        self.bill_laundry_times.setSuffix(" lần")
        self.bill_laundry_amount = QDoubleSpinBox()
        self.bill_laundry_amount.setMaximum(9_999_999)
        self.bill_laundry_amount.setDecimals(0)
        self.bill_laundry_amount.setSuffix(" đ")
        self.bill_laundry_times.valueChanged.connect(self._auto_fill_laundry_amount)

        laundry_row = QHBoxLayout()
        laundry_row.addWidget(self.bill_laundry_times)
        laundry_row.addWidget(QLabel(" hoặc số tiền:"))
        laundry_row.addWidget(self.bill_laundry_amount)
        laundry_widget = QWidget()
        laundry_widget.setLayout(laundry_row)

        # Other expenses
        self.bill_other_amount = QDoubleSpinBox()
        self.bill_other_amount.setMaximum(9_999_999)
        self.bill_other_amount.setDecimals(0)
        self.bill_other_amount.setSuffix(" đ")
        self.bill_other_note = QLineEdit()
        self.bill_other_note.setPlaceholderText("Ghi chú (VD: Sửa cửa)")

        inputs_form.addRow("Tiền Phòng:", self.bill_room_fee_lbl)
        inputs_form.addRow("Tiền Điện:", self.bill_elec_lbl)
        inputs_form.addRow("Tiền Nước:", self.bill_water_lbl)
        inputs_form.addRow("Tiền Giặt:", laundry_widget)
        inputs_form.addRow("Chi Phí Khác:", self.bill_other_amount)
        inputs_form.addRow("Ghi Chú Khác:", self.bill_other_note)

        btn_save_bill_data = QPushButton("💾 Lưu Giặt & Chi Phí Khác")
        btn_save_bill_data.clicked.connect(self._save_bill_extras)
        inputs_form.addRow(btn_save_bill_data)

        content.addWidget(inputs_group, 1)

        # Right: bill preview
        preview_group = QGroupBox("🧾 Hóa Đơn")
        preview_v = QVBoxLayout(preview_group)
        self.bill_preview = QTextEdit()
        self.bill_preview.setReadOnly(True)
        self.bill_preview.setFontFamily("Courier New")
        self.bill_preview.setFontPointSize(10)
        self.bill_preview.setPlaceholderText("Nhấn 'Tính Hóa Đơn' để xem kết quả...")
        preview_v.addWidget(self.bill_preview)

        btn_print = QPushButton("🖨️ In Hóa Đơn (Copy)")
        btn_print.clicked.connect(self._print_bill)
        preview_v.addWidget(btn_print)

        content.addWidget(preview_group, 1)
        layout.addLayout(content)

        self._refresh_bill_room_combo()
        return widget

    def _refresh_bill_room_combo(self):
        self.bill_room_combo.clear()
        rooms = self.db.get_all_rooms()
        for r in rooms:
            self.bill_room_combo.addItem(r["name"], r["id"])

    def _auto_fill_laundry_amount(self):
        times = self.bill_laundry_times.value()
        price = float(self.db.get_setting("laundry_price", 20000))
        self.bill_laundry_amount.setValue(times * price)

    def _calculate_bill(self):
        room_id = self.bill_room_combo.currentData()
        month = self.bill_month_combo.currentData()
        if not room_id:
            QMessageBox.warning(self, "Thông báo", "Vui lòng chọn phòng!")
            return

        data = self.db.get_bill_data(room_id, month)
        if not data:
            QMessageBox.warning(self, "Lỗi", "Không tìm thấy thông tin phòng!")
            return

        # Update labels
        self.bill_room_fee_lbl.setText(fmt_money(data["room_fee"]))

        if data["elec_usage"] > 0 or data["elec_fee"] > 0:
            self.bill_elec_lbl.setText(
                f"{fmt_money(data['elec_fee'])}  "
                f"({data['elec_usage']:.0f} kWh × {fmt_money(data['elec_price'])})"
            )
        else:
            self.bill_elec_lbl.setText("0đ  (chưa nhập số điện)")

        if data["water_usage"] > 0 or data["water_fee"] > 0:
            self.bill_water_lbl.setText(
                f"{fmt_money(data['water_fee'])}  "
                f"({data['water_usage']:.0f} khối × {fmt_money(data['water_price'])})"
            )
        else:
            self.bill_water_lbl.setText("0đ  (chưa nhập số nước)")

        # Load laundry & other from DB
        laundry = self.db.get_laundry(room_id, month)
        if laundry:
            self.bill_laundry_times.setValue(laundry["times"])
            self.bill_laundry_amount.setValue(laundry["amount"])
        else:
            self.bill_laundry_times.setValue(0)
            self.bill_laundry_amount.setValue(0)

        other = self.db.get_other_expense(room_id, month)
        if other:
            self.bill_other_amount.setValue(other["amount"])
            self.bill_other_note.setText(other["note"] or "")
        else:
            self.bill_other_amount.setValue(0)
            self.bill_other_note.setText("")

        # Re-read laundry/other from spinboxes (may have just been set)
        laundry_fee = self.bill_laundry_amount.value()
        other_fee = self.bill_other_amount.value()
        other_note = self.bill_other_note.text().strip()

        total = data["room_fee"] + data["elec_fee"] + data["water_fee"] + laundry_fee + other_fee

        # Generate bill text
        self._current_bill_data = {**data, "laundry_fee": laundry_fee, "other_fee": other_fee,
                                   "other_note": other_note, "total": total}
        self._render_bill_preview(self._current_bill_data)

    def _render_bill_preview(self, data):
        month_display = fmt_month_display(data["month"])
        lines = [
            "━━━━━━━━━━━━━━━━━━━━━━",
            "    HÓA ĐƠN TIỀN PHÒNG",
            "━━━━━━━━━━━━━━━━━━━━━━",
            "",
            f"Phòng: {data['room_name']}",
            f"Tháng: {month_display}",
            "",
            "━━━ CHI TIẾT ━━━",
            f"Tiền Phòng:   {fmt_money(data['room_fee'])}",
        ]

        if data.get("elec_fee", 0) > 0:
            lines.append(
                f"Tiền Điện:    {fmt_money(data['elec_fee'])}"
                f"  ({data['elec_usage']:.0f} kWh × {fmt_money(data['elec_price'])})"
            )
        else:
            lines.append("Tiền Điện:    0đ")

        if data.get("water_fee", 0) > 0:
            lines.append(
                f"Tiền Nước:    {fmt_money(data['water_fee'])}"
                f"  ({data['water_usage']:.0f} khối × {fmt_money(data['water_price'])})"
            )
        else:
            lines.append("Tiền Nước:    0đ")

        if data.get("laundry_fee", 0) > 0:
            lines.append(f"Tiền Giặt:    {fmt_money(data['laundry_fee'])}")
        else:
            lines.append("Tiền Giặt:    0đ")

        if data.get("other_fee", 0) > 0:
            note_str = f"  (Ghi chú: {data['other_note']})" if data.get("other_note") else ""
            lines.append(f"Chi Phí Khác: {fmt_money(data['other_fee'])}{note_str}")
        else:
            lines.append("Chi Phí Khác: 0đ")

        lines += [
            "",
            "━━━ TỔNG ━━━",
            f"TỔNG HÓA ĐƠN: {fmt_money(data['total'])}",
            "",
            "━━━━━━━━━━━━━━━━━━━━━━",
        ]
        self.bill_preview.setPlainText("\n".join(lines))

    def _save_bill_extras(self):
        room_id = self.bill_room_combo.currentData()
        month = self.bill_month_combo.currentData()
        if not room_id:
            QMessageBox.warning(self, "Thông báo", "Vui lòng chọn phòng!")
            return
        laundry_times = self.bill_laundry_times.value()
        laundry_amount = self.bill_laundry_amount.value()
        other_amount = self.bill_other_amount.value()
        other_note = self.bill_other_note.text().strip()

        self.db.save_laundry(room_id, month, laundry_times, laundry_amount)
        if other_amount > 0 or other_note:
            self.db.save_other_expense(room_id, month, other_amount, other_note)
        QMessageBox.information(self, "✅ Thành công", "Đã lưu tiền giặt và chi phí khác!")
        self._calculate_bill()

    def _print_bill(self):
        if not self.bill_preview.toPlainText():
            QMessageBox.warning(self, "Thông báo", "Vui lòng tính hóa đơn trước!")
            return
        clipboard = QApplication.clipboard()
        clipboard.setText(self.bill_preview.toPlainText())
        QMessageBox.information(self, "✅ Đã Copy", "Hóa đơn đã được copy vào clipboard!\nBạn có thể dán (Ctrl+V) vào Zalo, Word, hoặc máy in.")

    # ==================== TAB XUẤT EXCEL ====================
    def _create_export_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)

        info_lbl = QLabel(
            "📋 Xuất danh sách khách trọ đang ở ra file Excel.\n"
            "File phù hợp để khai báo tạm trú với công an."
        )
        info_lbl.setWordWrap(True)
        info_lbl.setAlignment(Qt.AlignCenter)
        f = QFont()
        f.setPointSize(10)
        info_lbl.setFont(f)
        layout.addWidget(info_lbl)
        layout.addSpacing(10)

        btn_export = QPushButton("📊 Xuất File Excel")
        btn_export.setMinimumHeight(45)
        f2 = QFont()
        f2.setPointSize(11)
        f2.setBold(True)
        btn_export.setFont(f2)
        btn_export.clicked.connect(self._export_excel)
        layout.addWidget(btn_export)
        layout.addSpacing(10)

        # Preview table
        preview_group = QGroupBox("👁️ Xem Trước Danh Sách")
        preview_v = QVBoxLayout(preview_group)

        btn_refresh_preview = QPushButton("🔄 Tải Danh Sách")
        btn_refresh_preview.clicked.connect(self._refresh_export_preview)
        preview_v.addWidget(btn_refresh_preview)

        self.export_preview_table = QTableWidget()
        self.export_preview_table.setColumnCount(6)
        self.export_preview_table.setHorizontalHeaderLabels(
            ["Phòng", "Tên", "CCCD", "Ngày Sinh", "Quê Quán", "SĐT"]
        )
        self.export_preview_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.export_preview_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.export_preview_table.setAlternatingRowColors(True)
        preview_v.addWidget(self.export_preview_table)
        layout.addWidget(preview_group)

        self._refresh_export_preview()
        return widget

    def _refresh_export_preview(self):
        tenants = self.db.get_all_active_tenants()
        self.export_preview_table.setRowCount(len(tenants))
        for row, t in enumerate(tenants):
            vals = [
                t["room_name"],
                t["name"],
                t["cccd"] or "",
                t["dob"] or "",
                t["hometown"] or "",
                t["phone"] or "",
            ]
            for col, v in enumerate(vals):
                self.export_preview_table.setItem(row, col, QTableWidgetItem(v))

    def _export_excel(self):
        if not _OPENPYXL_AVAILABLE:
            QMessageBox.critical(self, "❌ Lỗi", "Cần cài đặt openpyxl!\nChạy: pip install openpyxl")
            return

        tenants = self.db.get_all_active_tenants()
        if not tenants:
            QMessageBox.warning(self, "Thông báo", "Không có khách trọ nào đang ở!")
            return

        today_str = datetime.now().strftime("%Y-%m-%d")
        default_name = f"DanhSachKhachTro_{today_str}.xlsx"

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Lưu File Excel", default_name, "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Danh Sách Khách Trọ"

        # Header style
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = XlFont(bold=True, color="FFFFFF", size=11)
        header_align = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = ["Phòng", "Tên", "CCCD", "Ngày Sinh", "Quê Quán", "SĐT"]
        col_widths = [10, 25, 18, 14, 25, 15]

        # Write header row
        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = border
            ws.column_dimensions[cell.column_letter].width = width

        ws.row_dimensions[1].height = 22

        # Write data rows
        alt_fill = PatternFill("solid", fgColor="DCE6F1")
        for row_idx, t in enumerate(tenants, 2):
            vals = [
                t["room_name"],
                t["name"],
                t["cccd"] or "",
                t["dob"] or "",
                t["hometown"] or "",
                t["phone"] or "",
            ]
            for col_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = Alignment(vertical="center")
                cell.border = border
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        # Add title row at top
        ws.insert_rows(1)
        title_cell = ws.cell(row=1, column=1,
                             value=f"DANH SÁCH KHÁCH TRỌ - {datetime.now().strftime('%d/%m/%Y')}")
        title_cell.font = XlFont(bold=True, size=13)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        ws.row_dimensions[1].height = 25

        wb.save(file_path)
        QMessageBox.information(
            self, "✅ Xuất thành công",
            f"Đã xuất file Excel:\n{file_path}\n\nTổng: {len(tenants)} khách trọ"
        )

    # ==================== TAB CÀI ĐẶT ====================
    def _create_settings_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setAlignment(Qt.AlignTop)

        title_lbl = QLabel("⚙️ Cài Đặt Giá Mặc Định")
        f = QFont()
        f.setPointSize(12)
        f.setBold(True)
        title_lbl.setFont(f)
        title_lbl.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_lbl)
        layout.addSpacing(20)

        settings_group = QGroupBox("Giá Cả")
        form = QFormLayout(settings_group)
        form.setSpacing(12)

        self.setting_elec_price = QDoubleSpinBox()
        self.setting_elec_price.setMaximum(99_999)
        self.setting_elec_price.setDecimals(0)
        self.setting_elec_price.setSuffix(" đ/kWh")
        self.setting_elec_price.setMinimumWidth(180)

        self.setting_water_price = QDoubleSpinBox()
        self.setting_water_price.setMaximum(999_999)
        self.setting_water_price.setDecimals(0)
        self.setting_water_price.setSuffix(" đ/khối")

        self.setting_laundry_price = QDoubleSpinBox()
        self.setting_laundry_price.setMaximum(999_999)
        self.setting_laundry_price.setDecimals(0)
        self.setting_laundry_price.setSuffix(" đ/lần")

        # Load current values
        self.setting_elec_price.setValue(float(self.db.get_setting("electricity_price", 3500)))
        self.setting_water_price.setValue(float(self.db.get_setting("water_price", 15000)))
        self.setting_laundry_price.setValue(float(self.db.get_setting("laundry_price", 20000)))

        form.addRow("Giá Điện:", self.setting_elec_price)
        form.addRow("Giá Nước:", self.setting_water_price)
        form.addRow("Giá Giặt:", self.setting_laundry_price)

        layout.addWidget(settings_group)
        layout.addSpacing(20)

        btn_save = QPushButton("💾 Lưu Cài Đặt")
        btn_save.setMinimumHeight(40)
        f2 = QFont()
        f2.setPointSize(11)
        btn_save.setFont(f2)
        btn_save.clicked.connect(self._save_settings)
        layout.addWidget(btn_save)

        note_lbl = QLabel(
            "💡 Lưu ý:\n"
            "• Giá điện: đơn giá mỗi kWh\n"
            "• Giá nước: đơn giá mỗi khối (m³)\n"
            "• Giá giặt: đơn giá mỗi lần giặt\n"
            "• Thay đổi giá sẽ áp dụng cho hóa đơn mới"
        )
        note_lbl.setWordWrap(True)
        note_lbl.setStyleSheet("color: #555; padding: 10px;")
        layout.addWidget(note_lbl)

        return widget

    def _save_settings(self):
        self.db.set_setting("electricity_price", self.setting_elec_price.value())
        self.db.set_setting("water_price", self.setting_water_price.value())
        self.db.set_setting("laundry_price", self.setting_laundry_price.value())
        QMessageBox.information(self, "✅ Thành công", "Đã lưu cài đặt giá!")

    # ==================== DASHBOARD & NOTIFICATIONS ====================
    def _refresh_dashboard(self):
        stats = self.db.get_statistics()
        self.lbl_total_rooms.setText(f"🏘️ Tổng phòng: {stats['total_rooms']}")
        self.lbl_empty_rooms.setText(f"🔵 Phòng trống: {stats['empty_rooms']}")
        self.lbl_total_tenants.setText(f"👥 Tổng khách trọ: {stats['total_tenants']}")

    def _check_billing_notifications(self):
        """Kiểm tra và thông báo phòng tới hạn đóng tiền hôm nay"""
        rooms_due = self.db.get_rooms_due_today()
        if rooms_due:
            room_names = ", ".join([r["name"] for r in rooms_due])
            today = datetime.now()
            msg = QMessageBox(self)
            msg.setWindowTitle("🔔 Thông Báo Hạn Đóng Tiền")
            msg.setIcon(QMessageBox.Information)
            msg.setText(
                f"Hôm nay là ngày {today.day}/{today.month}/{today.year}\n\n"
                f"Các phòng tới hạn đóng tiền:\n\n"
                f"  📌 {room_names}\n\n"
                f"Vui lòng nhắc nhở khách trọ!"
            )
            msg.addButton("OK, Tôi Biết", QMessageBox.AcceptRole)
            msg.exec_()

    def _refresh_all_combos(self):
        """Refresh tất cả các combobox chọn phòng (dùng sau khi thêm/xóa phòng)"""
        self._refresh_util_room_combo()
        self._refresh_bill_room_combo()
        self._refresh_export_preview()


def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    window = RoomManagementApp()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
